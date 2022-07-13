import math
from openpyxl import load_workbook

def stats_bitcoin(): 
    ''' retorna las fechas y porcentajes del 2021, formato tupla(fecha(str), porc(float))
    '''

    #los datos obtenidos son el porcentaje de variacion diaria del Bitcoin 
    # desde 01/01/2021 hasta 01/01/2022
    
    #arreglo de datos de bitcoin
    datos = []

    #formato: fecha, porcentaje de variacion
    #excel = load_workbook(filename= r"E:\matia\Documents\Facultad\SIM\TPs\TP2\simulacion\TP2\STATS_BIT.xlsx")
    excel = load_workbook(
        filename=r"TP2\STATS_BIT.xlsx")  # Ruta del exel donde se extraen los datos

    #selecciona la unica hoja del excel
    hoja_seleccionada = excel['STATS_BIT']
    fechas = hoja_seleccionada['A']
    variaciones = hoja_seleccionada['G']

    #flag para omitir la cabecera de los datos
    es_cabecera = True

    #recupera los strings con los datos, todavia sin separar
    for i in range(len(fechas)):
        if es_cabecera:
            es_cabecera = False
            continue
        
        datos.append(
            (fechas[i].value,
            float(variaciones[i].value
            .replace("%", "")   #se le quita el %
            .replace(",", "."))))   #se pasa el , a .

    return datos

if __name__ == "__main__":
    print(stats_bitcoin)